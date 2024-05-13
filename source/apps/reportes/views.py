from django.http import HttpResponse
from django.shortcuts import render

from django.views import View
from django.views.generic import TemplateView
from datetime import datetime
from collections import Counter
from django.db.models import Count
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from openpyxl import Workbook
from apps.mutual.forms import FormulariosFinalizados
from apps.mutual.models import DeclaracionJurada, Mutual, Periodo
from django.db.models import Sum
from django.utils import formats

class reporteMutualDeclaracionesJuradasView(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    template_name = 'reporte_mutuales_declaraciones.html'
    login_url = "/login/"
    permission_required = "empleadospublicos.permission_empleado_publico"
    
    def get_graph_mutual(self, mutual):
        declaraciones = DeclaracionJurada.objects.filter(mutual=mutual)
        
        reclamos = []
        prestamos = []
        categorias = []
        if declaraciones.exists():
            for declaracion in declaraciones:
                fecha = declaracion.periodo.mes_anio.strftime('%Y-%m')  # Formatear la fecha como una cadena
                categorias.append(fecha)
                
                importe_reclamo = float(declaracion.detalles.filter(tipo='R').first().importe) if declaracion.detalles.filter(tipo='R').first() else 0
                importe_prestamo = float(declaracion.detalles.filter(tipo='P').first().importe) if declaracion.detalles.filter(tipo='P').first() else 0
                
                reclamos.append(importe_reclamo)
                prestamos.append(importe_prestamo)
                
        return categorias, reclamos, prestamos
        

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mutual_id = self.kwargs.get('mutual_id')
        mutual = get_object_or_404(Mutual, pk=mutual_id)

        categorias, reclamos, prestamos = self.get_graph_mutual(mutual)

        context['titulo'] = f"MUTUAL {mutual.nombre} Y SUS DECLARACIONES JURADAS PRESENTADAS"
        context['categorias'] = categorias
        context['reclamos'] = reclamos
        context['prestamos'] = prestamos
        context['mutual_id'] = mutual_id  # Pass mutual_id to JavaScript
        return context
    
from openpyxl.styles import Font, Alignment, PatternFill    

def reporte_periodo_dj(request):
    if request.method == 'GET':
        # periodosFinalizados = Periodo.objects.filter(fecha_fin__isnull = False)
        periodosFinalizados = FormulariosFinalizados()
        return render(request, 'reporte_periodo_dj.html', {'periodos': periodosFinalizados})
    
    if request.method == 'POST':
        try:
            id_periodo = request.POST.get('periodos')
            periodo = Periodo.objects.get(pk=id_periodo)
            declaraciones = DeclaracionJurada.objects.filter(periodo=periodo)
            declaraciones =  declaraciones.order_by('mutual__alias')
        
            wb = Workbook()
            ws = wb.active
            ws.title = "DECLARACIONES"
            
            # Estilos para el encabezado
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="77dd77", end_color="77dd77", fill_type="solid")
            
            ws.append(["Periodo" , "N° Declaraciones"])
            ws.append([periodo.mes_anio, declaraciones.count()])
            ws.append([""])
            ws.append([""])
            # Añadir encabezados
            headers = ["Mutual", "Tipo Archivo", "Concepto", "Importe", "N° Registros"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
            
            for cell in ws[5]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
            
            # Estilos para los datos
            data_font = Font(size=11)
            
            # Agregar datos
            for declaracion in declaraciones:
                mutual = declaracion.mutual.alias
                for detalle in declaracion.detalles.all():
                    tipo = "Prestamo"
                    if detalle.tipo != 'P':
                        tipo = "Reclamo"
                        
                    row = [mutual, tipo, detalle.concepto, detalle.importe, detalle.total_registros]
                    ws.append(row)
                    
                    
                # Separador entre declaraciones
                # ws.append([""])
                
            # Configurar ancho de columnas
            column_widths = [20, 15,10,15, 10]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64+i)].width = width

            
            
            # ws2 = wb.create_sheet(title="NO DECLARADOS")
            # ws2['A1'] = "Contenido de la otra hoja"
            
            
            
            
            
            # Crear la respuesta HTTP con el contenido del archivo Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            extension = ".xlsx"
            nombre = periodo.mes_anio.strftime('%Y%m')+ "_ReportePeriodo" + extension
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'
            

            # Guardar el contenido del libro de trabajo en la respuesta HTTP
            wb.save(response)

            return response
        
        except Periodo.DoesNotExist:
            print("No existe periodo")